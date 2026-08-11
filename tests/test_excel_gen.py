import openpyxl
import os
import re

sample_plan_data = {
    "title": "Programa Comunitário de Formação e Marcenaria",
    "proponent": "Associação dos Moradores",
    "institution": "Edital Rio Doce",
    "items": [
        {
            "itemGroup": "Serviços Especializados (PF e PJ)",
            "natureza": "outros serviços de terceiros",
            "descricao": "Profissional responsável pela execução das oficinas de MARCENARIA",
            "unid": "horas",
            "qtde": 936,
            "valorPrevisto": 69.0,
            "valorTotal": 64584.0,
            "atividade": 1
        },
        {
            "itemGroup": "Máquinas e Equipamentos",
            "natureza": "material permanente",
            "descricao": "DESENGROSSO DEWALT DW733 220V PORTATIL",
            "unid": "unidade",
            "qtde": 1,
            "valorPrevisto": 3800.0,
            "valorTotal": 3800.0,
            "atividade": 1
        }
    ]
}

def parse_num(val, fallback=0.0):
    if val is None: return fallback
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip()
    s = re.sub(r'[^\d.,-]', '', s)
    if not s: return fallback
    if ',' in s and '.' in s:
        if s.find('.') < s.find(','):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return fallback

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = "Planilha Orçamentária"

ws1.cell(row=1, column=1, value="EDITAL RIO DOCE - PLANILHA ORÇAMENTÁRIA DO PROJETO")
ws1.cell(row=2, column=1, value="NOME DO PROJETO:")
ws1.cell(row=2, column=3, value=sample_plan_data["title"])
ws1.cell(row=3, column=1, value="PROPONENTE:")
ws1.cell(row=3, column=3, value=sample_plan_data["proponent"])

headers1 = ["ITEM", "NATUREZA", "DESCRIÇÃO DO ITEM / SERVIÇO", "UNID", "QTDE", "VALOR PREVISTO", "VALOR TOTAL", "ATIVIDADE"]
for col_idx, h in enumerate(headers1, start=1):
    ws1.cell(row=8, column=col_idx, value=h)

start_r = 9
for idx, it in enumerate(sample_plan_data["items"]):
    r = start_r + idx
    qtd = parse_num(it.get('qtde'), 1.0)
    v_unit = parse_num(it.get('valorPrevisto'), 0.0)

    ws1.cell(row=r, column=1, value=it.get('itemGroup'))
    ws1.cell(row=r, column=2, value=it.get('natureza'))
    ws1.cell(row=r, column=3, value=it.get('descricao'))
    ws1.cell(row=r, column=4, value=it.get('unid'))
    ws1.cell(row=r, column=5, value=qtd)
    ws1.cell(row=r, column=6, value=v_unit)
    ws1.cell(row=r, column=7, value=f"=E{r}*F{r}")
    ws1.cell(row=r, column=8, value=it.get('atividade'))

tot_r = start_r + len(sample_plan_data["items"])
ws1.cell(row=tot_r, column=1, value="TOTAL DA META 1")
ws1.cell(row=tot_r, column=7, value=f"=SUM(G{start_r}:G{tot_r-1})")

tot_geral_r = tot_r + 1
ws1.cell(row=tot_geral_r, column=1, value="TOTAL GERAL DO PROJETO")
ws1.cell(row=tot_geral_r, column=7, value=f"=G{tot_r}")

out_file = "test_rio_doce.xlsx"
wb.save(out_file)

wb_read = openpyxl.load_workbook(out_file, data_only=False)
ws_check = wb_read["Planilha Orçamentária"]

print("Row 9 (Item 1) Item:", ws_check["A9"].value)
print("Row 9 (Item 1) Qtde:", ws_check["E9"].value)
print("Row 9 (Item 1) ValorPrevisto:", ws_check["F9"].value)
print("Row 9 (Item 1) ValorTotal Formula:", ws_check["G9"].value)
print("Row 11 Meta Total Formula:", ws_check["G11"].value)
print("Row 12 Total Geral Formula:", ws_check["G12"].value)

assert ws_check["A9"].value == "Serviços Especializados (PF e PJ)"
assert ws_check["E9"].value == 936
assert ws_check["F9"].value == 69.0
assert ws_check["G9"].value == "=E9*F9"
assert ws_check["G11"].value == "=SUM(G9:G10)"
assert ws_check["G12"].value == "=G11"

print("SUCCESS: EDITAL RIO DOCE LAYOUT CHECKS PASSED PERFECTLY!")

if os.path.exists(out_file):
    os.remove(out_file)
