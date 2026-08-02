#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Servidor Customizado HTTP para EditalAudit AI
Suporta servir arquivos estáticos e atua como Proxy para carregamento de links de editais.
"""

import os
import json
import urllib.request
import urllib.parse
import urllib.error
import re
import html
import io
import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from html.parser import HTMLParser
from services.api import LLMGateway, DocumentRetriever
from services.skills.anki_exporter import create_anki_apkg_zip

# ReportLab imports at top-level
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Global divider helper for ReportLab reports
def get_divider():
    line = Table([['']], colWidths=[487], rowHeights=[1])
    line.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#e2e8f0')),
        ('PADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    return line

gateway = LLMGateway()

def safe_encode_cp1252(s):
    b = bytearray()
    for char in s:
        cp = ord(char)
        if 0x80 <= cp <= 0x9f:
            try:
                b.extend(char.encode('cp1252'))
            except UnicodeEncodeError:
                b.append(cp)
        else:
            b.extend(char.encode('cp1252'))
    return bytes(b)

def fix_double_encoded_utf8(text):
    if not isinstance(text, str) or not text:
        return text
    
    if any(c in text for c in ('Ã', 'Â', 'â', 'Ê', 'Ô')):
        for enc in ('cp1252', 'latin-1'):
            try:
                if enc == 'cp1252':
                    return safe_encode_cp1252(text).decode('utf-8')
                else:
                    return text.encode(enc).decode('utf-8')
            except (UnicodeEncodeError, UnicodeDecodeError):
                pass
            
    def _sub_fix(match):
        for enc in ('cp1252', 'latin-1'):
            try:
                if enc == 'cp1252':
                    return safe_encode_cp1252(match.group(0)).decode('utf-8')
                else:
                    return match.group(0).encode(enc).decode('utf-8')
            except (UnicodeEncodeError, UnicodeDecodeError):
                pass
        return match.group(0)

    # In cp1252/latin-1 double-encoding:
    # 2-byte UTF-8 starts with 0xc2-0xdf, followed by continuation byte
    # 3-byte UTF-8 starts with 0xe0-0xef, followed by two continuation bytes
    pattern = re.compile(r'[\u00c2-\u00df].|[\u00e0-\u00ef].{2}')
    text = pattern.sub(_sub_fix, text)
    return text

def format_ptbr_currency(val):
    if isinstance(val, (int, float)):
        return f"R$ {val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    if isinstance(val, str) and val.strip():
        if val.startswith("R$"):
            return val
        try:
            clean_str = val.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
            float_val = float(clean_str)
            return f"R$ {float_val:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except ValueError:
            return val
    return str(val) if val is not None else "R$ 0,00"

def clean_html_tags(temp_text):
    if not temp_text:
        return ""
    # 1. Strip HTML comments
    temp_text = re.sub(r'<!--[\s\S]*?-->', '', temp_text)
    # 2. Strip style and script tags and contents
    temp_text = re.sub(r'<style[^>]*>[\s\S]*?</style>', '', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'<script[^>]*>[\s\S]*?</script>', '', temp_text, flags=re.IGNORECASE)
    # 3. Headers to bold + br
    temp_text = re.sub(r'<h[1-6][^>]*>(.*?)</h[1-6]>', r'<br/><b>\1</b><br/>', temp_text, flags=re.DOTALL | re.IGNORECASE)
    # 4. List items to bullets
    temp_text = re.sub(r'<li[^>]*>(.*?)</li>', r'• \1<br/>', temp_text, flags=re.DOTALL | re.IGNORECASE)
    temp_text = re.sub(r'</?(?:ul|ol)[^>]*>', r'<br/>', temp_text, flags=re.IGNORECASE)
    
    # 5. Table cells and headers
    temp_text = re.sub(r'<th[^>]*>(.*?)</th>', r' | <b>\1</b> ', temp_text, flags=re.DOTALL | re.IGNORECASE)
    temp_text = re.sub(r'<td[^>]*>(.*?)td>', r' | \1 ', temp_text, flags=re.DOTALL | re.IGNORECASE)
    temp_text = re.sub(r'<tr[^>]*>', '', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'</tr>', '<br/>', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'</?(?:table|tbody|thead|tfoot)[^>]*>', '<br/>', temp_text, flags=re.IGNORECASE)
    
    # 6. Strong / em to b / i
    temp_text = re.sub(r'<strong[^>]*>', '<b>', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'</strong>', '</b>', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'<em[^>]*>', '<i>', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'</em>', '</i>', temp_text, flags=re.IGNORECASE)
    temp_text = re.sub(r'</?(?:p|div|section|article|header|footer)[^>]*>', r'<br/>', temp_text, flags=re.IGNORECASE)
    
    # 7. Strip any other tag except ReportLab allowed: b, i, u, sub, sup, font, a, br
    allowed_prefixes = ('<b', '</b', '<i', '</i', '<u', '</u', '<sub', '</sub', '<sup', '</sup', '<font', '</font', '<a', '</a', '<br', '</br')
    def strip_unallowed(m):
        tag = m.group(0)
        tag_lower = tag.lower()
        if any(tag_lower.startswith(prefix) for prefix in allowed_prefixes):
            return tag
        return ''
        
    temp_text = re.sub(r'<[^>]+>', strip_unallowed, temp_text)
    return temp_text

def append_html_content_to_story(html_content, story, body_style, h2_style):
    if not html_content:
        return

    clean_html = re.sub(r'<!--[\s\S]*?-->', '', html_content)
    clean_html = re.sub(r'<style[^>]*>[\s\S]*?</style>', '', clean_html, flags=re.IGNORECASE)
    clean_html = re.sub(r'<script[^>]*>[\s\S]*?</script>', '', clean_html, flags=re.IGNORECASE)

    table_pattern = re.compile(r'(<table[\s\S]*?>[\s\S]*?</table>)', re.IGNORECASE)
    blocks = table_pattern.split(clean_html)

    for block in blocks:
        block_str = block.strip()
        if not block_str:
            continue

        if block_str.lower().startswith('<table'):
            parser = HTMLTableParser()
            parser.feed(block_str)
            rows = parser.rows
            if rows:
                N = max(len(r) for r in rows)
                col_widths = [487.0 / N] * N
                table_content = []
                for row in rows:
                    row_cells = []
                    for cell in row:
                        cell_text = make_reportlab_safe(cell["text"])
                        try:
                            if cell["is_header"]:
                                cell_p = Paragraph(f"<b>{cell_text}</b>", ParagraphStyle('ThCustom', parent=body_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#0f172a')))
                            else:
                                cell_p = Paragraph(cell_text, body_style)
                        except Exception:
                            esc_txt = html.escape(re.sub(r'<[^>]+>', '', cell_text))
                            cell_p = Paragraph(esc_txt, body_style)
                        row_cells.append(cell_p)
                    while len(row_cells) < N:
                        row_cells.append(Paragraph("", body_style))
                    table_content.append(row_cells)

                report_table = Table(table_content, colWidths=col_widths)
                t_style = TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                    ('PADDING', (0,0), (-1,-1), 5),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ])
                for r_idx in range(1, len(table_content)):
                    if r_idx % 2 == 1:
                        t_style.add('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#f8fafc'))
                report_table.setStyle(t_style)
                story.append(Spacer(1, 4))
                story.append(report_table)
                story.append(Spacer(1, 6))
        else:
            temp_text = clean_html_tags(block_str)
            parts = re.split(r'<br/>|<br>', temp_text)
            for part in parts:
                clean_part = part.strip()
                if clean_part:
                    safe_part = make_reportlab_safe(clean_part)
                    if safe_part.strip():
                        if safe_part.startswith('<b>') and safe_part.endswith('</b>') and len(safe_part) < 100:
                            story.append(Paragraph(safe_part, h2_style))
                        else:
                            try:
                                story.append(Paragraph(safe_part, body_style))
                            except Exception as pe:
                                plain_text = re.sub(r'<[^>]+>', '', safe_part)
                                story.append(Paragraph(html.escape(plain_text), body_style))


def make_reportlab_safe(text):
    if not text:
        return ""
    text = str(text)
    
    # Fix double-encoded UTF-8 first
    text = fix_double_encoded_utf8(text)
    
    # Replace common MS Word / Unicode smart quotes, dashes, bullets and special characters
    replacements = {
        '\u201c': '"', '\u201d': '"', '\u201e': '"', '\u201f': '"', '\u2033': '"', '\u2036': '"',
        '\u2018': "'", '\u2019': "'", '\u201a': "'", '\u201b': "'", '\u2032': "'", '\u2035': "'",
        '\u2012': '-', '\u2013': '-', '\u2014': '-', '\u2015': '-',
        '\u2022': '*', '\u2023': '*', '\u2043': '*', '\u204c': '*', '\u204d': '*', '\u2219': '*', '\u25aa': '*', '\u25ab': '*',
        '\u2026': '...',
        '\u00a0': ' ',
        '\u200b': '', '\u200c': '', '\u200d': '', '\ufeff': '',
    }
    for orig, rep in replacements.items():
        text = text.replace(orig, rep)
        
    # Decode HTML entities if any
    text = html.unescape(text)
    # Escape HTML special characters (< and >) safely without turning quotes into &quot;
    text = html.escape(text, quote=False)
    
    # Restore allowed ReportLab tags
    text = text.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    text = text.replace("&lt;i&gt;", "<i>").replace("&lt;/i&gt;", "</i>")
    text = text.replace("&lt;u&gt;", "<u>").replace("&lt;/u&gt;", "</u>")
    text = text.replace("&lt;sub&gt;", "<sub>").replace("&lt;/sub&gt;", "</sub>")
    text = text.replace("&lt;sup&gt;", "<sup>").replace("&lt;/sup&gt;", "</sup>")
    text = text.replace("&lt;br&gt;", "<br/>").replace("&lt;br/&gt;", "<br/>").replace("&lt;br /&gt;", "<br/>")
    
    # Restore font tags: &lt;font (.*?)&gt; -> <font \1>
    text = re.sub(r'&lt;font\s+(.*?)&gt;', r'<font \1>', text, flags=re.IGNORECASE)
    text = text.replace("&lt;/font&gt;", "</font>").replace("&lt;/FONT&gt;", "</font>")
    
    # Restore a tags: &lt;a\s+(.*?)&gt; -> <a \1>
    text = re.sub(r'&lt;a\s+(.*?)&gt;', r'<a \1>', text, flags=re.IGNORECASE)
    text = text.replace("&lt;/a&gt;", "</a>").replace("&lt;/A&gt;", "</a>")
    
    return text

PORT = 8085

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.2.1 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
]

def search_ddg(query):
    import random
    ua = random.choice(USER_AGENTS)
    headers = {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7"
    }
    
    # Tentativa 1: DuckDuckGo HTML
    url = "https://html.duckduckgo.com/html/?" + urllib.parse.urlencode({"q": query})
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=12) as response:
            html_content = response.read().decode('utf-8', errors='ignore')
            results = []
            pattern = re.compile(r'<a[^>]+class="[^"]*result__a[^"]*"[^>]*href="([^"]+)"[^>]*>([\s\S]*?)</a>')
            matches = pattern.findall(html_content)
            
            for href, title in matches:
                title_clean = re.sub(r'<[^>]+>', '', title).strip()
                title_clean = html.unescape(title_clean)
                if "/l/?kh=" in href or "uddg=" in href:
                    parsed_url = urllib.parse.urlparse(href)
                    qs = urllib.parse.parse_qs(parsed_url.query)
                    if 'uddg' in qs:
                        href = qs['uddg'][0]
                results.append({"title": title_clean, "url": href, "snippet": ""})
            
            snippet_pattern = re.compile(r'<a class="result__snippet"[^>]*>([\s\S]*?)</a>')
            snippets = snippet_pattern.findall(html_content)
            for i, snip in enumerate(snippets):
                if i < len(results):
                    snippet_clean = re.sub(r'<[^>]+>', '', snip).strip()
                    snippet_clean = html.unescape(snippet_clean)
                    results[i]["snippet"] = snippet_clean
                    
            if results:
                return results[:15]
    except Exception as e:
        print(f"[SEARCH][WARN] DuckDuckGo HTML falhou: {e}. Tentando fallback Lite...")

    # Tentativa 2: DuckDuckGo Lite Fallback
    try:
        lite_url = "https://lite.duckduckgo.com/lite/?" + urllib.parse.urlencode({"q": query})
        req_lite = urllib.request.Request(lite_url, headers=headers)
        with urllib.request.urlopen(req_lite, timeout=12) as response:
            html_content = response.read().decode('utf-8', errors='ignore')
            results = []
            link_matches = re.findall(r'<a[^>]+href="([^"]+)"[^>]*>([\s\S]*?)</a>', html_content)
            for href, title in link_matches:
                title_clean = html.unescape(re.sub(r'<[^>]+>', '', title).strip())
                if href.startswith('http') and len(title_clean) > 5:
                    results.append({"title": title_clean, "url": href, "snippet": "Diretriz / Edital de fomento cultural público."})
            return results[:15]
    except Exception as e2:
        print(f"[SEARCH][ERROR] Fallback DuckDuckGo Lite falhou: {e2}")
        return []

def extract_document_links(html_content, base_url):
    link_pattern = re.compile(r'<a[^>]+href="([^"]+)"[^>]*>([\s\S]*?)</a>', re.IGNORECASE)
    matches = link_pattern.findall(html_content)
    
    links = []
    seen_urls = set()
    
    doc_extensions = ('.pdf', '.docx', '.doc', '.txt', '.odt')
    
    for href, text in matches:
        href = href.strip()
        href = href.replace('&amp;', '&')
        full_url = urllib.parse.urljoin(base_url, href)
        
        parsed = urllib.parse.urlparse(full_url)
        if parsed.scheme not in ('http', 'https'):
            continue
            
        text_clean = re.sub(r'<[^>]+>', '', text).strip()
        text_clean = " ".join(text_clean.split())
        text_clean = text_clean.replace('&amp;', '&').replace('&quot;', '"').replace('&#39;', "'")
        
        if not text_clean:
            text_clean = os.path.basename(parsed.path) or "Documento"
            
        is_doc = any(parsed.path.lower().endswith(ext) for ext in doc_extensions)
        contains_keywords = any(kw in text_clean.lower() or kw in parsed.path.lower() for kw in ['edital', 'regulamento', 'anexo', 'chamada', 'retificacao', 'cronograma', 'contrato'])
        
        if (is_doc or contains_keywords) and full_url not in seen_urls:
            seen_urls.add(full_url)
            links.append({
                "name": text_clean,
                "url": full_url,
                "is_direct_doc": is_doc
            })
            
    return links

class HTMLTextExtractor(HTMLParser):
    def __init__(self):
        super().__init__()
        self.text = []
        self.ignored_tags = set()

    def handle_starttag(self, tag, attrs):
        if tag in ["script", "style", "head", "title", "meta", "link"]:
            self.ignored_tags.add(tag)

    def handle_endtag(self, tag):
        if tag in ["script", "style", "head", "title", "meta", "link"]:
            self.ignored_tags.discard(tag)

    def handle_data(self, data):
        if not self.ignored_tags:
            self.text.append(data)

    def get_clean_text(self):
        full_text = " ".join(self.text)
        return " ".join(full_text.split())


class HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self.current_row = []
        self.current_cell = []
        self.in_cell = False
        self.is_header = False

    def handle_starttag(self, tag, attrs):
        if tag == 'tr':
            self.current_row = []
        elif tag in ['td', 'th']:
            self.in_cell = True
            self.is_header = (tag == 'th')
            self.current_cell = []

    def handle_endtag(self, tag):
        if tag == 'tr':
            if self.current_row:
                self.rows.append(self.current_row)
        elif tag in ['td', 'th']:
            self.in_cell = False
            cell_text = "".join(self.current_cell).strip()
            self.current_row.append({"text": cell_text, "is_header": self.is_header})

    def handle_data(self, data):
        if self.in_cell:
            self.current_cell.append(data)


class CustomHTTPRequestHandler(SimpleHTTPRequestHandler):

    def do_GET(self):
        if self.path == '/favicon.ico':
            self.send_response(204)
            self.end_headers()
            return
        super().do_GET()

    def send_header(self, keyword, value):
        if keyword.lower() == 'content-type':
            if any(text_type in value.lower() for text_type in ['text/html', 'text/javascript', 'application/javascript', 'text/css', 'application/json']):
                if 'charset' not in value.lower():
                    value += '; charset=utf-8'
        super().send_header(keyword, value)

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def do_POST(self):
        if self.path == '/api/fetch-url':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            
            try:
                data = json.loads(post_data.decode('utf-8'))
                url = data.get('url')
                
                if not url:
                    self.send_json_response(400, {"error": "URL ausente no corpo da requisição."})
                    return

                # Realiza a requisição ao link do edital
                req = urllib.request.Request(
                    url, 
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                    }
                )
                
                with urllib.request.urlopen(req, timeout=12) as response:
                    content = response.read()
                    content_type = response.info().get_content_type()
                    
                    # Se for PDF ou Word (DOCX/DOC), retorna os bytes brutos para o browser processar
                    if 'application/pdf' in content_type or 'application/vnd.openxmlformats' in content_type or 'application/msword' in content_type:
                        self.send_response(200)
                        self.send_header('Content-Type', content_type)
                        self.send_header('Content-Length', str(len(content)))
                        self.end_headers()
                        self.wfile.write(content)
                        return
                    else:
                        # Se for HTML ou texto, decodifica e extrai o texto limpo
                        raw_charset = response.info().get_content_charset()
                        if not raw_charset:
                            try:
                                html_content = content.decode('utf-8')
                            except Exception:
                                html_content = content.decode('latin1', errors='replace')
                        else:
                            try:
                                html_content = content.decode(raw_charset)
                            except Exception:
                                html_content = content.decode('utf-8', errors='replace')
                        
                        # Extrai texto limpo usando parser embutido
                        parser = HTMLTextExtractor()
                        parser.feed(html_content)
                        clean_text = parser.get_clean_text()
                        
                        response_data = {
                            "text": clean_text,
                            "content_type": content_type
                        }
                        
                        self.send_json_response(200, response_data)
                        return

            except urllib.error.HTTPError as e:
                self.send_json_response(500, {"error": f"Erro HTTP {e.code} ao obter conteúdo da URL."})
            except urllib.error.URLError as e:
                self.send_json_response(500, {"error": f"Falha de conexão ou URL inválida: {str(e.reason)}"})
            except Exception as e:
                self.send_json_response(500, {"error": f"Erro inesperado no servidor proxy: {str(e)}"})
        
        elif self.path == '/api/search-web-editais':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                query = data.get('query')
                if not query:
                    self.send_json_response(400, {"error": "Termo de busca (query) ausente."})
                    return
                
                results = search_ddg(query)
                self.send_json_response(200, {"results": results})
            except Exception as e:
                self.send_json_response(500, {"error": f"Erro ao pesquisar: {str(e)}"})
                
        elif self.path == '/api/parse-portal-page':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                url = data.get('url')
                if not url:
                    self.send_json_response(400, {"error": "URL ausente."})
                    return
                
                req = urllib.request.Request(
                    url,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                    }
                )
                
                with urllib.request.urlopen(req, timeout=12) as response:
                    content_type = response.info().get_content_type()
                    
                    doc_extensions = ('application/pdf', 'application/vnd.openxmlformats', 'application/msword')
                    is_direct_doc = any(ext in content_type for ext in doc_extensions) or any(url.lower().endswith(ext) for ext in ('.pdf', '.docx', '.doc'))
                    
                    if is_direct_doc:
                        filename = os.path.basename(urllib.parse.urlparse(url).path) or "Edital.pdf"
                        self.send_json_response(200, {
                            "type": "document",
                            "url": url,
                            "name": filename
                        })
                        return
                    else:
                        content = response.read()
                        raw_charset = response.info().get_content_charset()
                        if not raw_charset:
                            try:
                                html_content = content.decode('utf-8')
                            except Exception:
                                html_content = content.decode('latin1', errors='replace')
                        else:
                            try:
                                html_content = content.decode(raw_charset)
                            except Exception:
                                html_content = content.decode('utf-8', errors='replace')
                        
                        links = extract_document_links(html_content, url)
                        self.send_json_response(200, {
                            "type": "portal",
                            "links": links
                        })
                        return
            except Exception as e:
                self.send_json_response(500, {"error": f"Erro ao analisar portal: {str(e)}"})
        
        elif self.path == '/api/generate-audit-pdf':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                project_title = str(data.get('project_title') or 'Projeto Cultural')
                institution = str(data.get('institution') or 'Não Especificada')
                proponent = str(data.get('proponent') or 'Não Especificado')
                budget = str(data.get('budget') or '0')
                score = str(data.get('score') or '0')
                nota_tecnica = str(data.get('nota_tecnica') or '0')
                nota_priorizacao = str(data.get('nota_priorizacao') or '0')
                relatorio_analitico = str(data.get('relatorio_analitico') or '')
                criterios = data.get('criterios', [])
                ajustes = data.get('ajustes', [])
                alertas = data.get('alertas', [])
                
                # Normalize values to empty lists if they are None/null
                if criterios is None:
                    criterios = []
                if ajustes is None:
                    ajustes = []
                if alertas is None:
                    alertas = []
                
                # Imports cleaned up (now top-level)
                
                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=A4,
                    leftMargin=54,
                    rightMargin=54,
                    topMargin=54,
                    bottomMargin=54
                )
                
                styles = getSampleStyleSheet()
                
                title_style = ParagraphStyle(
                    'DocTitle',
                    parent=styles['Heading1'],
                    fontName='Helvetica-Bold',
                    fontSize=18,
                    leading=22,
                    textColor=colors.HexColor('#1e1b4b'),
                    spaceAfter=6
                )
                subtitle_style = ParagraphStyle(
                    'DocSubtitle',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=10,
                    leading=13,
                    textColor=colors.HexColor('#4f46e5'),
                    spaceAfter=15
                )
                h2_style = ParagraphStyle(
                    'SectionHeader',
                    parent=styles['Heading2'],
                    fontName='Helvetica-Bold',
                    fontSize=13,
                    leading=16,
                    textColor=colors.HexColor('#0f172a'),
                    spaceBefore=14,
                    spaceAfter=6,
                    keepWithNext=True
                )
                body_style = ParagraphStyle(
                    'BodyTextCustom',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=9.5,
                    leading=13.5,
                    textColor=colors.HexColor('#334155'),
                    spaceAfter=6
                )
                score_style = ParagraphStyle(
                    'ScoreStyle',
                    parent=styles['Normal'],
                    fontName='Helvetica-Bold',
                    fontSize=11,
                    leading=14,
                    textColor=colors.HexColor('#4f46e5')
                )
                
                story = []
                
                # Make header elements reportlab safe
                safe_title = make_reportlab_safe("RELATÓRIO DE AUDITORIA GERAL DE COMPLIANCE")
                safe_proj_title = make_reportlab_safe(project_title)
                safe_proponent = make_reportlab_safe(proponent)
                safe_institution = make_reportlab_safe(institution)
                
                story.append(Paragraph(safe_title, title_style))
                story.append(Paragraph(f"Projeto: <b>{safe_proj_title}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Proponente: {safe_proponent}", subtitle_style))
                story.append(Spacer(1, 10))
                
                # Divider helper cleaned up (uses global)
                
                # Executive Summary Table
                summary_data = [
                    [
                        Paragraph("<b>Fomento/Órgão:</b>", body_style), Paragraph(safe_institution, body_style),
                        Paragraph("<b>Orçamento:</b>", body_style), Paragraph(f"R$ {budget}", body_style)
                    ],
                    [
                        Paragraph("<b>Nota Técnica:</b>", body_style), Paragraph(f"{nota_tecnica} pts", body_style),
                        Paragraph("<b>Nota Priorização:</b>", body_style), Paragraph(f"{nota_priorizacao} pts", body_style)
                    ]
                ]
                summary_table = Table(summary_data, colWidths=[100, 140, 110, 137])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
                    ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                    ('PADDING', (0,0), (-1,-1), 6),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                
                story.append(Paragraph("<b>Sumário Executivo</b>", h2_style))
                story.append(summary_table)
                story.append(Spacer(1, 10))
                
                # Calculate max_score
                try:
                    max_score = sum(int(crit.get('nota_maxima', 20) or 20) for crit in criterios) if (criterios and len(criterios) > 0) else 100
                except Exception as sum_e:
                    print(f"Error summing max_score: {sum_e}")
                    max_score = 100
                
                # Score Table & Progress Bar
                score_html = f"<b>Nota Geral de Compliance:</b> <font color='#4f46e5' size=14><b>{score} / {max_score}</b></font>"
                score_table_data = [[Paragraph(score_html, score_style)]]
                score_table = Table(score_table_data, colWidths=[487])
                score_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f1f5f9')),
                    ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
                    ('PADDING', (0,0), (-1,-1), 10),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ]))
                story.append(score_table)
                story.append(Spacer(1, 6))
                
                # Visual Progress Bar
                score_num = 0
                try:
                    score_num = float(score)
                except:
                    pass
                percent = min(100.0, max(0.0, (score_num / float(max_score)) * 100)) if max_score > 0 else 0
                width_filled = max(1, int(487 * (percent / 100.0)))
                width_empty = max(1, 487 - width_filled)
                bar_color = colors.HexColor('#10b981') if percent >= 70 else (colors.HexColor('#f59e0b') if percent >= 50 else colors.HexColor('#ef4444'))
                
                progress_table = Table([['', '']], colWidths=[width_filled, width_empty], rowHeights=[8])
                progress_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (0,0), bar_color),
                    ('BACKGROUND', (1,0), (1,0), colors.HexColor('#e2e8f0')),
                    ('PADDING', (0,0), (-1,-1), 0),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 0),
                    ('TOPPADDING', (0,0), (-1,-1), 0),
                ]))
                story.append(progress_table)
                story.append(Spacer(1, 15))
                story.append(get_divider())
                story.append(Spacer(1, 5))
                
                if relatorio_analitico:
                    story.append(Paragraph("Parecer Técnico Descritivo da Auditoria", h2_style))
                    append_html_content_to_story(relatorio_analitico, story, body_style, h2_style)
                    story.append(Spacer(1, 10))
                
                story.append(Paragraph("Quesitos Analisados (Instrução Normativa MinC)", h2_style))
                for crit in criterios:
                    crit_name = make_reportlab_safe(crit.get('criterio', 'Critério'))
                    try:
                        nota_atrib = int(crit.get('nota_atribuida', 0) or 0)
                    except:
                        nota_atrib = 0
                    try:
                        nota_max = int(crit.get('nota_maxima', 25) or 25)
                    except:
                        nota_max = 25
                    just = make_reportlab_safe(crit.get('justificativa', ''))
                    
                    ratio = float(nota_atrib) / float(nota_max) if nota_max > 0 else 0
                    crit_color = '#10b981' if ratio >= 0.8 else ('#d97706' if ratio >= 0.5 else '#ef4444')
                    crit_title = f"<font color='{crit_color}'><b>{crit_name} ({nota_atrib}/{nota_max} pts)</b></font>"
                    try:
                        story.append(Paragraph(crit_title, ParagraphStyle('CritHeader', parent=body_style, fontName='Helvetica-Bold')))
                        story.append(Paragraph(just, body_style))
                    except Exception as pe:
                        print(f"ReportLab criteria rendering error: {pe}")
                        story.append(Paragraph(html.escape(f"{crit.get('criterio', 'Critério')} ({crit.get('nota_atribuida', 0)}/{crit.get('nota_maxima', 25)} pts)"), ParagraphStyle('CritHeader', parent=body_style, fontName='Helvetica-Bold')))
                        story.append(Paragraph(html.escape(crit.get('justificativa', '')), body_style))
                    story.append(Spacer(1, 6))
                
                story.append(Spacer(1, 10))
                
                if ajustes:
                    story.append(Paragraph("Ajustes Operacionais Recomendados", h2_style))
                    table_data = [[
                        Paragraph("<b>Alteração Sugerida</b>", ParagraphStyle('Th', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)),
                        Paragraph("<b>Fator de Impacto</b>", ParagraphStyle('Th', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white))
                    ]]
                    for a in ajustes:
                        alt_text = make_reportlab_safe(a.get('alteracao', ''))
                        fator_text = make_reportlab_safe(a.get('fator', ''))
                        try:
                            alt_p = Paragraph(alt_text, body_style)
                            fator_p = Paragraph(fator_text, body_style)
                        except Exception as pe:
                            print(f"ReportLab adjustments table Paragraph error: {pe}")
                            alt_p = Paragraph(html.escape(a.get('alteracao', '')), body_style)
                            fator_p = Paragraph(html.escape(a.get('fator', '')), body_style)
                        table_data.append([alt_p, fator_p])
                    
                    ajustes_table = Table(table_data, colWidths=[337, 150])
                    ajustes_table.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                        ('PADDING', (0,0), (-1,-1), 6),
                        ('VALIGN', (0,0), (-1,-1), 'TOP'),
                        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8fafc')),
                    ]))
                    story.append(ajustes_table)
                    story.append(Spacer(1, 15))
                
                if alertas:
                    story.append(Paragraph("Alertas Críticos Jurídicos & Inconsistências", h2_style))
                    for alert in alertas:
                        al_type = alert.get('tipo', 'Alerta')
                        desc = alert.get('descricao', '')
                        sug = alert.get('sugestao', '')
                        nivel = alert.get('nivel', 'MEDIA').upper()
                        
                        border_color = colors.HexColor('#ef4444') if nivel == 'ALTA' else (colors.HexColor('#f59e0b') if nivel == 'MEDIA' else colors.HexColor('#10b981'))
                        bg_color = colors.HexColor('#fef2f2') if nivel == 'ALTA' else (colors.HexColor('#fffbeb') if nivel == 'MEDIA' else colors.HexColor('#f0fdf4'))
                        
                        safe_type = make_reportlab_safe(al_type)
                        safe_desc = make_reportlab_safe(desc)
                        safe_sug = make_reportlab_safe(sug)
                        
                        alert_html = f"<b>[{nivel}] {safe_type}:</b> {safe_desc}<br/><i>Recomendação: {safe_sug}</i>"
                        
                        try:
                            alert_table_data = [[Paragraph(alert_html, body_style)]]
                        except Exception as pe:
                            print(f"ReportLab alerts Paragraph error: {pe}")
                            esc_type = html.escape(al_type)
                            esc_desc = html.escape(desc)
                            esc_sug = html.escape(sug)
                            alert_html_fallback = f"<b>[{nivel}] {esc_type}:</b> {esc_desc}<br/><i>Recomendação: {esc_sug}</i>"
                            alert_table_data = [[Paragraph(alert_html_fallback, body_style)]]
                        
                        alert_table = Table(alert_table_data, colWidths=[487])
                        alert_table.setStyle(TableStyle([
                            ('BACKGROUND', (0,0), (-1,-1), bg_color),
                            ('BOX', (0,0), (-1,-1), 1, border_color),
                            ('LINELEFT', (0,0), (-1,-1), 4, border_color),
                            ('PADDING', (0,0), (-1,-1), 8),
                            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ]))
                        story.append(alert_table)
                        story.append(Spacer(1, 8))
                
                story.append(Spacer(1, 10))
                
                # Disclaimer
                disclaimer_style = ParagraphStyle(
                    'Disclaimer',
                    parent=styles['Normal'],
                    fontName='Helvetica-Oblique',
                    fontSize=8,
                    leading=10,
                    textColor=colors.HexColor('#64748b'),
                    spaceBefore=15
                )
                story.append(Paragraph("Este relatório é uma auditoria preliminar baseada em simulação por inteligência artificial estruturada e leitura estática de conformidade do edital. As notas e recomendações não garantem aprovação do projeto perante a comissão oficial.", disclaimer_style))
                
                def add_footer(canvas, doc):
                    canvas.saveState()
                    canvas.setFont('Helvetica', 8)
                    canvas.setFillColor(colors.HexColor('#64748b'))
                    import datetime
                    date_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                    canvas.drawString(54, 30, f"Gerado por EditalAudit AI em {date_str}")
                    canvas.drawRightString(A4[0] - 54, 30, f"Página {doc.page}")
                    canvas.restoreState()
                    
                doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
                pdf_bytes = pdf_buffer.getvalue()
                pdf_buffer.close()
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/pdf')
                import unicodedata
                filename_clean = ''.join(c for c in unicodedata.normalize('NFD', project_title) if unicodedata.category(c) != 'Mn')
                filename_clean = re.sub(r'[^a-zA-Z0-9]', '_', filename_clean)
                filename_clean = re.sub(r'_+', '_', filename_clean).strip('_')
                if not filename_clean or filename_clean.lower() == 'titulo_do_projeto_cultural':
                    filename_clean = "Projeto_Cultural"
                self.send_header('Content-Disposition', f'attachment; filename="Laudo_Auditoria_Compliance_{filename_clean}.pdf"')
                self.send_header('Content-Length', str(len(pdf_bytes)))
                self.end_headers()
                self.wfile.write(pdf_bytes)
                return
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.send_json_response(500, {"error": f"Erro ao gerar PDF da auditoria: {str(e)}"})

        elif self.path == '/api/generate-revisor-report-pdf':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                project_title = str(data.get('project_title') or 'Projeto Cultural')
                institution = str(data.get('institution') or 'Não Especificada')
                report_content = str(data.get('report_content') or '')
                
                # Imports cleaned up (now top-level)
                
                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=A4,
                    leftMargin=54,
                    rightMargin=54,
                    topMargin=54,
                    bottomMargin=54
                )
                
                styles = getSampleStyleSheet()
                
                title_style = ParagraphStyle(
                    'DocTitle',
                    parent=styles['Heading1'],
                    fontName='Helvetica-Bold',
                    fontSize=18,
                    leading=22,
                    textColor=colors.HexColor('#1e1b4b'),
                    spaceAfter=6
                )
                subtitle_style = ParagraphStyle(
                    'DocSubtitle',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=10,
                    leading=13,
                    textColor=colors.HexColor('#4f46e5'),
                    spaceAfter=15
                )
                h2_style = ParagraphStyle(
                    'SectionHeader',
                    parent=styles['Heading2'],
                    fontName='Helvetica-Bold',
                    fontSize=13,
                    leading=16,
                    textColor=colors.HexColor('#0f172a'),
                    spaceBefore=14,
                    spaceAfter=6,
                    keepWithNext=True
                )
                body_style = ParagraphStyle(
                    'BodyTextCustom',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=9.5,
                    leading=13.5,
                    textColor=colors.HexColor('#334155'),
                    spaceAfter=6
                )
                
                story = []
                
                # Make header elements reportlab safe
                safe_title = make_reportlab_safe("RELATÓRIO CONSOLIDADO DE REVISÃO E PLANO DE AÇÃO")
                safe_proj_title = make_reportlab_safe(project_title)
                safe_institution = make_reportlab_safe(institution)
                
                story.append(Paragraph(safe_title, title_style))
                story.append(Paragraph(f"Projeto: <b>{safe_proj_title}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Órgão: {safe_institution}", subtitle_style))
                story.append(Spacer(1, 10))
                
                # Divider helper cleaned up (uses global)
                
                story.append(get_divider())
                story.append(Spacer(1, 10))
                
                # Parse report_content HTML tags and structure
                append_html_content_to_story(report_content, story, body_style, h2_style)
                
                story.append(Spacer(1, 15))
                story.append(get_divider())
                
                # Disclaimer
                disclaimer_style = ParagraphStyle(
                    'Disclaimer',
                    parent=styles['Normal'],
                    fontName='Helvetica-Oblique',
                    fontSize=8,
                    leading=10,
                    textColor=colors.HexColor('#64748b'),
                    spaceBefore=15
                )
                story.append(Paragraph("Este documento é um relatório consolidado de revisão analítica e não constitui aprovação ou homologação oficial da proposta.", disclaimer_style))
                
                def add_revisor_footer(canvas, doc):
                    canvas.saveState()
                    canvas.setFont('Helvetica', 8)
                    canvas.setFillColor(colors.HexColor('#64748b'))
                    import datetime
                    date_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                    canvas.drawString(54, 30, f"Gerado por EditalAudit AI em {date_str}")
                    canvas.drawRightString(A4[0] - 54, 30, f"Página {doc.page}")
                    canvas.restoreState()
                    
                doc.build(story, onFirstPage=add_revisor_footer, onLaterPages=add_revisor_footer)
                pdf_bytes = pdf_buffer.getvalue()
                pdf_buffer.close()
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/pdf')
                import unicodedata
                filename_clean = ''.join(c for c in unicodedata.normalize('NFD', project_title) if unicodedata.category(c) != 'Mn')
                filename_clean = re.sub(r'[^a-zA-Z0-9]', '_', filename_clean)
                filename_clean = re.sub(r'_+', '_', filename_clean).strip('_')
                if not filename_clean or filename_clean.lower() == 'titulo_do_projeto_cultural':
                    filename_clean = "Projeto_Cultural"
                self.send_header('Content-Disposition', f'attachment; filename="Relatorio_Detalhado_Revisor_{filename_clean}.pdf"')
                self.send_header('Content-Length', str(len(pdf_bytes)))
                self.end_headers()
                self.wfile.write(pdf_bytes)
                return
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.send_json_response(500, {"error": f"Erro ao gerar PDF da revisão: {str(e)}"})

        elif self.path == '/api/generate-finance-pdf':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                project_title = str(data.get('project_title') or 'Projeto Cultural')
                proponent = str(data.get('proponent') or 'Não Especificado')
                institution = str(data.get('institution') or 'Não Especificada')
                budget = str(data.get('budget') or '0')
                table_html = str(data.get('table_html') or '')
                
                page_format = landscape(A4)
                printable_width = page_format[0] - 72 # 36pt margins on landscape

                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=page_format,
                    leftMargin=36,
                    rightMargin=36,
                    topMargin=36,
                    bottomMargin=36
                )
                
                styles = getSampleStyleSheet()
                
                title_style = ParagraphStyle(
                    'DocTitle',
                    parent=styles['Heading1'],
                    fontName='Helvetica-Bold',
                    fontSize=16,
                    leading=20,
                    textColor=colors.HexColor('#1e1b4b'),
                    spaceAfter=4
                )
                subtitle_style = ParagraphStyle(
                    'DocSubtitle',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=9,
                    leading=12,
                    textColor=colors.HexColor('#4f46e5'),
                    spaceAfter=12
                )
                body_style = ParagraphStyle(
                    'TableBodyText',
                    parent=styles['Normal'],
                    fontName='Helvetica',
                    fontSize=7.0,
                    leading=9.0,
                    textColor=colors.HexColor('#334155')
                )
                
                story = []
                
                header_title = f"{institution.upper()} - PLANILHA ORÇAMENTÁRIA DO PROJETO" if institution and institution.lower() != 'edital' else "PLANILHA ORÇAMENTÁRIA DO PROJETO"
                safe_title = make_reportlab_safe(header_title)
                safe_proj_title = make_reportlab_safe(project_title)
                safe_proponent = make_reportlab_safe(proponent)
                safe_institution = make_reportlab_safe(institution)
                
                story.append(Paragraph(safe_title, title_style))
                story.append(Paragraph(f"Projeto: <b>{safe_proj_title}</b> &nbsp;&nbsp;|&nbsp;&nbsp; Proponente: {safe_proponent}", subtitle_style))
                story.append(Spacer(1, 8))
                
                # Create executive summary table for finance header
                summary_data = [
                    [
                        Paragraph("<b>Fomento / Edital:</b>", body_style), Paragraph(safe_institution, body_style),
                        Paragraph("<b>Orçamento Previsto:</b>", body_style), Paragraph(f"R$ {budget}", body_style)
                    ]
                ]
                summary_table = Table(summary_data, colWidths=[90, printable_width*0.4, 110, printable_width*0.35])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
                    ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                    ('PADDING', (0,0), (-1,-1), 6),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 10))

                items = data.get('items', [])
                grand_subtotal = data.get('grandTotalSubtotal', 0)
                grand_impostos = data.get('grandTotalImpostos', 0)
                grand_geral = data.get('grandTotalGeral', 0)

                # Se items não foi passado, usar parser legados de HTML como fallback
                if not items and table_html:
                    parser = HTMLTableParser()
                    parser.feed(table_html)
                    rows = parser.rows
                    items = []
                    for r in rows:
                        if len(r) >= 8 and not r[0]["is_header"]:
                            items.append({
                                "itemGroup": r[0]["text"],
                                "natureza": r[1]["text"],
                                "descricao": r[2]["text"],
                                "unid": r[3]["text"],
                                "qtde": r[4]["text"],
                                "valorPrevisto": r[5]["text"],
                                "valorTotal": r[6]["text"],
                                "atividade": r[7]["text"]
                            })

                if items:
                    headers = ["ITEM / CATEGORIA", "NATUREZA", "DESCRIÇÃO DO ITEM / SERVIÇO", "UNID.", "QTDE", "VALOR PREVISTO (R$)", "VALOR TOTAL (R$)", "ATIVIDADE"]
                    weights = [0.18, 0.18, 0.28, 0.07, 0.05, 0.10, 0.09, 0.05]
                    col_widths = [printable_width * w for w in weights]
                    
                    table_content = []
                    # Add Header Row
                    header_cells = [Paragraph(f"<b>{make_reportlab_safe(h)}</b>", ParagraphStyle('ThFinance', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)) for h in headers]
                    table_content.append(header_cells)
                    
                    # Add Item Rows
                    for idx, it in enumerate(items):
                        val_unit = format_ptbr_currency(it.get('valorUnit', it.get('valorPrevisto', 0)))
                        tot_ger = format_ptbr_currency(it.get('total', it.get('valorTotal', 0)))
                        item_cat = str(it.get('rubrica', it.get('itemGroup', 'Serviços Especializados')))
                        nat_str = str(it.get('destino', it.get('natureza', 'outros serviços de terceiros')))
                        desc_str = str(it.get('item', it.get('descricao', 'Descrição do Serviço')))
                        unid_str = str(it.get('unidade', it.get('unid', 'unidade')))
                        qtd_str = str(it.get('qtd', it.get('qtde', 1)))
                        ativ_str = str(it.get('atividade', (idx % 3) + 1))

                        row_cells = [
                            Paragraph(make_reportlab_safe(item_cat), body_style),
                            Paragraph(make_reportlab_safe(nat_str), body_style),
                            Paragraph(f"<b>{make_reportlab_safe(desc_str)}</b>", body_style),
                            Paragraph(make_reportlab_safe(unid_str), body_style),
                            Paragraph(make_reportlab_safe(qtd_str), body_style),
                            Paragraph(make_reportlab_safe(val_unit), body_style),
                            Paragraph(f"<b>{make_reportlab_safe(tot_ger)}</b>", body_style),
                            Paragraph(make_reportlab_safe(ativ_str), body_style)
                        ]
                        table_content.append(row_cells)

                    # Add Total Row
                    tot_ger_str = f"R$ {grand_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') if isinstance(grand_geral, (int, float)) else str(grand_geral)

                    total_row_cells = [
                        Paragraph("<b>TOTAL GERAL DO PROJETO:</b>", ParagraphStyle('TotLbl', parent=body_style, fontName='Helvetica-Bold', alignment=2)),
                        Paragraph("", body_style), Paragraph("", body_style), Paragraph("", body_style),
                        Paragraph("", body_style), Paragraph("", body_style),
                        Paragraph(f"<b>{tot_ger_str}</b>", ParagraphStyle('TotVal', parent=body_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#15803d'))),
                        Paragraph("", body_style)
                    ]
                    table_content.append(total_row_cells)
                        
                    finance_table = Table(table_content, colWidths=col_widths, repeatRows=1)
                    t_style = TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563eb')),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                        ('PADDING', (0,0), (-1,-1), 3),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ('SPAN', (0, -1), (5, -1)),
                        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1e40af')),
                        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                    ])
                    # Alternating row colors
                    for r_idx in range(1, len(table_content) - 1):
                        if r_idx % 2 == 0:
                            t_style.add('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#f8fafc'))
                    finance_table.setStyle(t_style)
                    story.append(finance_table)
                    story.append(Spacer(1, 12))

                # Rider Técnico Section (if rider_items passed)
                rider_items = data.get('rider_items', [])
                if rider_items and isinstance(rider_items, list):
                    h2_style = ParagraphStyle(
                        'SectionHeader',
                        parent=styles['Heading2'],
                        fontName='Helvetica-Bold',
                        fontSize=12,
                        leading=15,
                        textColor=colors.HexColor('#4f46e5'),
                        spaceBefore=10,
                        spaceAfter=6
                    )
                    story.append(Paragraph("DETALHAMENTO DO RIDER TÉCNICO & MAPA DE EQUIPAMENTOS", h2_style))
                    
                    rider_table_data = [
                        [
                            Paragraph("<b>Categoria</b>", ParagraphStyle('ThRider', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)),
                            Paragraph("<b>Equipamento</b>", ParagraphStyle('ThRider', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)),
                            Paragraph("<b>Modelo Específico / Especificação</b>", ParagraphStyle('ThRider', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)),
                            Paragraph("<b>Diárias / Qtd</b>", ParagraphStyle('ThRider', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)),
                            Paragraph("<b>Fornecedor Previsto</b>", ParagraphStyle('ThRider', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white)),
                            Paragraph("<b>Requisito de Palco</b>", ParagraphStyle('ThRider', parent=body_style, fontName='Helvetica-Bold', textColor=colors.white))
                        ]
                    ]
                    
                    for rd in rider_items:
                        rider_table_data.append([
                            Paragraph(make_reportlab_safe(rd.get('categoria', '')), body_style),
                            Paragraph(make_reportlab_safe(rd.get('equipamento', '')), body_style),
                            Paragraph(make_reportlab_safe(rd.get('modeloEspecifico', '')), body_style),
                            Paragraph(make_reportlab_safe(rd.get('qtdDiarias', '')), body_style),
                            Paragraph(make_reportlab_safe(rd.get('fornecedorPrevisto', '')), body_style),
                            Paragraph(make_reportlab_safe(rd.get('requisitoPalco', '')), body_style)
                        ])
                        
                    r_widths = [printable_width * w for w in [0.15, 0.20, 0.28, 0.10, 0.15, 0.12]]
                    rider_table = Table(rider_table_data, colWidths=r_widths, repeatRows=1)
                    rt_style = TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6366f1')),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                        ('PADDING', (0,0), (-1,-1), 3),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ])
                    for r_idx in range(1, len(rider_table_data)):
                        if r_idx % 2 == 0:
                            rt_style.add('BACKGROUND', (0, r_idx), (-1, r_idx), colors.HexColor('#f8fafc'))
                    rider_table.setStyle(rt_style)
                    story.append(rider_table)
                    story.append(Spacer(1, 10))
                
                # Disclaimer
                disclaimer_style = ParagraphStyle(
                    'Disclaimer',
                    parent=styles['Normal'],
                    fontName='Helvetica-Oblique',
                    fontSize=8,
                    leading=10,
                    textColor=colors.HexColor('#64748b'),
                    spaceBefore=10
                )
                story.append(Paragraph("Este documento foi consolidado pelas 3 Etapas de Auditoria com base na legislação de fomento cultural (Lei Rouanet, Lei Aldir Blanc, IN MinC).", disclaimer_style))
                
                def add_finance_footer(canvas, doc):
                    canvas.saveState()
                    canvas.setFont('Helvetica', 8)
                    canvas.setFillColor(colors.HexColor('#64748b'))
                    import datetime
                    date_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                    canvas.drawString(54, 30, f"Gerado por EditalAudit AI em {date_str}")
                    canvas.drawRightString(A4[0] - 54, 30, f"Página {doc.page}")
                    canvas.restoreState()
                    
                doc.build(story, onFirstPage=add_finance_footer, onLaterPages=add_finance_footer)
                pdf_bytes = pdf_buffer.getvalue()
                pdf_buffer.close()
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/pdf')
                import unicodedata
                filename_clean = ''.join(c for c in unicodedata.normalize('NFD', project_title) if unicodedata.category(c) != 'Mn')
                filename_clean = re.sub(r'[^a-zA-Z0-9]', '_', filename_clean)
                filename_clean = re.sub(r'_+', '_', filename_clean).strip('_')
                if not filename_clean or filename_clean.lower() == 'titulo_do_projeto_cultural':
                    filename_clean = "Projeto_Cultural"
                self.send_header('Content-Disposition', f'attachment; filename="Planilha_Financeira_{filename_clean}.pdf"')
                self.send_header('Content-Length', str(len(pdf_bytes)))
                self.end_headers()
                self.wfile.write(pdf_bytes)
                return
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.send_json_response(500, {"error": f"Erro ao gerar PDF do financeiro: {str(e)}"})

        elif self.path == '/api/export-finance-xlsx':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                
                header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                rider_header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
                summary_header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
                timeline_header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
                
                title_font = Font(name="Segoe UI", size=14, bold=True, color="1E1B4B")
                sub_font = Font(name="Segoe UI", size=9, italic=True, color="475569")
                bold_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
                regular_font = Font(name="Segoe UI", size=10, color="1E293B")
                
                zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
                
                thin_side = Side(border_style="thin", color="CBD5E1")
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
                currency_fmt = '"R$" #,##0.00'
                qty_fmt = '#,##0'
                pct_fmt = '0.0%'

                def parse_num(val, fallback=0.0):
                    if val is None: return fallback
                    if isinstance(val, (int, float)): return float(val)
                    s = str(val).strip()
                    s = re.sub(r'[^\d.,-]', '', s)
                    if not s: return fallback
                    if ',' in s and '.' in s:
                        if s.find('.') < s.find(','):
                            s = s.replace('.', '').replace(',', '.')
                        else:
                            s = s.replace(',', '')
                    elif ',' in s:
                        s = s.replace(',', '.')
                    try:
                        return float(s)
                    except ValueError:
                        return fallback

                def clean_str(s):
                    if s is None: return ""
                    return fix_double_encoded_utf8(str(s)).strip()

                project_title = clean_str(data.get('title') or 'Projeto Cultural')
                proponent = clean_str(data.get('proponent') or 'Proponente')
                institution = clean_str(data.get('institution') or 'Edital')
                
                raw_items = data.get('items', [])
                items = [
                    it for it in raw_items 
                    if isinstance(it, dict) and "Subtotal" not in str(it.get('subtotal', '')) and "Item de Despesa" not in str(it.get('item', ''))
                ]
                rider_items = data.get('riderItems', [])

                # ABA 1: Planilha Orçamentária (Modelo de Referência Flexível para Editais)
                ws1 = wb.active
                ws1.title = "Planilha Orçamentária"
                
                # Cabeçalho Dinâmico de Identificação do Projeto
                header_title = f"{institution.upper()} - PLANILHA ORÇAMENTÁRIA DO PROJETO" if institution and institution.lower() != 'edital' else "PLANILHA ORÇAMENTÁRIA DO PROJETO"
                ws1.cell(row=1, column=1, value=header_title).font = title_font
                
                ws1.cell(row=2, column=1, value="NOME DO PROJETO:").font = bold_font
                ws1.cell(row=2, column=3, value=project_title).font = bold_font
                
                ws1.cell(row=3, column=1, value="PROPONENTE:").font = bold_font
                ws1.cell(row=3, column=3, value=proponent).font = regular_font
                
                ws1.cell(row=4, column=1, value="OBJETIVO GERAL:").font = bold_font
                ws1.cell(row=4, column=3, value=f"Execução integral das ações socioculturais conforme aprovação no {institution}.").font = regular_font
                
                # Banner Amarelo de Observação Normativa Generica de Fomento
                obs_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
                obs_font = Font(name="Segoe UI", size=8.5, italic=True, color="854D0E")
                obs_cell = ws1.cell(row=5, column=1, value="OBSERVAÇÃO NORMATIVA: Os valores apresentados foram dimensionados conforme pesquisa de mercado e limites de fomento, visando eficiência, transparência e rigor fiscal.")
                obs_cell.fill = obs_fill
                obs_cell.font = obs_font
                ws1.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)
                
                # Barra de Objetivo Específico & Meta
                oe_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                meta_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                header_text_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                
                oe_cell = ws1.cell(row=6, column=1, value="OBJETIVO ESPECÍFICO: OE 1 - REALIZAÇÃO E OPERACIONALIZAÇÃO INTEGRAL DO PROJETO")
                oe_cell.fill = oe_fill
                oe_cell.font = header_text_font
                ws1.merge_cells(start_row=6, start_column=1, end_row=6, end_column=8)
                
                meta_cell = ws1.cell(row=7, column=1, value="META: M1 - EXECUÇÃO DAS ATIVIDADES PRINCIPAIS, CONTRATAÇÃO DE EQUIPE E SUPRIMENTOS")
                meta_cell.fill = meta_fill
                meta_cell.font = header_text_font
                ws1.merge_cells(start_row=7, start_column=1, end_row=7, end_column=8)
                
                # Cabeçalho Oficial da Tabela (Colunas A até H)
                headers1 = [
                    "ITEM / CATEGORIA", "NATUREZA", "DESCRIÇÃO DO ITEM / SERVIÇO", 
                    "UNID", "QTDE", "VALOR PREVISTO (R$)", "VALOR TOTAL (R$)", "ATIVIDADE"
                ]
                
                for col_idx, h in enumerate(headers1, start=1):
                    c = ws1.cell(row=8, column=col_idx, value=h)
                    c.fill = meta_fill
                    c.font = header_text_font
                    c.alignment = Alignment(horizontal="center" if col_idx in (4, 5, 8) else ("right" if col_idx in (6, 7) else "left"), vertical="center")
                    c.border = thin_border
                
                start_r = 9
                for idx, it in enumerate(items):
                    r = start_r + idx
                    fill = zebra_fill if idx % 2 == 1 else None
                    
                    qtd = parse_num(it.get('qtd', it.get('qtde', 1)), 1.0)
                    v_unit = parse_num(it.get('valorUnit', it.get('valorPrevisto', 0)), 0.0)

                    item_cat = clean_str(it.get('rubrica', it.get('itemGroup', 'Serviços Especializados')))
                    natureza = clean_str(it.get('destino', it.get('natureza', 'outros serviços de terceiros')))
                    desc = clean_str(it.get('item', it.get('descricao', 'Descrição do Serviço')))
                    unid = clean_str(it.get('unidade', it.get('unid', 'unidade')))
                    ativ = idx % 3 + 1

                    row_vals = [
                        (item_cat, regular_font, None, "left"),
                        (natureza, regular_font, None, "left"),
                        (desc, bold_font, None, "left"),
                        (unid, regular_font, None, "center"),
                        (qtd, regular_font, qty_fmt, "center"),
                        (v_unit, regular_font, currency_fmt, "right"),
                        (f"=E{r}*F{r}", bold_font, currency_fmt, "right"),
                        (ativ, regular_font, None, "center")
                    ]
                    
                    for c_idx, (val, fn, num_fmt, align) in enumerate(row_vals, start=1):
                        c = ws1.cell(row=r, column=c_idx, value=val)
                        c.font = fn
                        if fill: c.fill = fill
                        c.border = thin_border
                        c.alignment = Alignment(horizontal=align, vertical="center")
                        if num_fmt: c.number_format = num_fmt

                tot_r = start_r + len(items)
                
                # Linha de Subtotal da Meta 1
                subtotal_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
                ws1.cell(row=tot_r, column=1, value="TOTAL DA META 1").font = bold_font
                ws1.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=6)
                ws1.cell(row=tot_r, column=1).alignment = Alignment(horizontal="right", vertical="center")
                
                sum_sub = ws1.cell(row=tot_r, column=7, value=f"=SUM(G{start_r}:G{tot_r-1})")
                sum_sub.font = bold_font; sum_sub.number_format = currency_fmt; sum_sub.border = thin_border; sum_sub.fill = subtotal_fill
                sum_sub.alignment = Alignment(horizontal="right", vertical="center")
                ws1.cell(row=tot_r, column=8, value="").border = thin_border
                
                # Linha de Total Geral do Projeto
                tot_geral_r = tot_r + 1
                ws1.cell(row=tot_geral_r, column=1, value="TOTAL GERAL DO PROJETO").font = header_text_font
                ws1.cell(row=tot_geral_r, column=1).fill = oe_fill
                ws1.merge_cells(start_row=tot_geral_r, start_column=1, end_row=tot_geral_r, end_column=6)
                ws1.cell(row=tot_geral_r, column=1).alignment = Alignment(horizontal="right", vertical="center")

                sum_tot = ws1.cell(row=tot_geral_r, column=7, value=f"=G{tot_r}")
                sum_tot.font = header_text_font; sum_tot.number_format = currency_fmt; sum_tot.border = thin_border; sum_tot.fill = oe_fill
                sum_tot.alignment = Alignment(horizontal="right", vertical="center")
                ws1.cell(row=tot_geral_r, column=8, value="").fill = oe_fill; ws1.cell(row=tot_geral_r, column=8).border = thin_border

                for col in ws1.columns:
                    max_l = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws1.column_dimensions[col_letter].width = max(max_l + 3, 14)

                # ABA 2: Rider Técnico & Equipamentos
                ws2 = wb.create_sheet(title="Rider Técnico & Equipamentos")
                ws2.cell(row=1, column=1, value="Detalhamento do Rider Técnico & Equipamentos de Palco").font = title_font
                ws2.cell(row=2, column=1, value=f"Especificações dos sistemas de som, iluminação e praticáveis para: {project_title}").font = sub_font
                
                headers2 = ["Categoria", "Equipamento / Estrutura", "Modelo Específico / Especificação", "Qtd / Diárias", "Fornecedor Previsto", "Requisito de Palco / ART"]
                for c_idx, h in enumerate(headers2, start=1):
                    c = ws2.cell(row=4, column=c_idx, value=h)
                    c.fill = rider_header_fill; c.font = header_font; c.border = thin_border
                    c.alignment = Alignment(horizontal="center" if c_idx == 4 else "left", vertical="center")

                for idx, rd in enumerate(rider_items):
                    r = 5 + idx
                    fill = zebra_fill if idx % 2 == 1 else None
                    r_vals = [
                        (rd.get('categoria', 'Geral'), bold_font),
                        (rd.get('equipamento', ''), bold_font),
                        (rd.get('modeloEspecifico', ''), regular_font),
                        (rd.get('qtdDiarias', '1'), regular_font),
                        (rd.get('fornecedorPrevisto', ''), regular_font),
                        (rd.get('requisitoPalco', ''), regular_font),
                    ]
                    for c_idx, (v, fn) in enumerate(r_vals, start=1):
                        c = ws2.cell(row=r, column=c_idx, value=v)
                        c.font = fn; c.border = thin_border
                        if fill: c.fill = fill
                        if c_idx == 4: c.alignment = Alignment(horizontal="center")

                for col in ws2.columns:
                    max_l = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws2.column_dimensions[col_letter].width = max(max_l + 3, 14)

                # ABA 3: Resumo Executivo & Memória de Cálculo
                ws3 = wb.create_sheet(title="Resumo & Memória de Cálculo")
                ws3.cell(row=1, column=1, value="Resumo Executivo, Limites Legais & Memória de Cálculo").font = title_font
                
                headers3 = ["Indicador Normativo", "Valor / Proporção Calculada", "Teto Legal do Edital", "Parecer de Conformidade Legal"]
                for c_idx, h in enumerate(headers3, start=1):
                    c = ws3.cell(row=3, column=c_idx, value=h)
                    c.fill = summary_header_fill; c.font = header_font; c.border = thin_border

                sum_rows = [
                    ("Orçamento Geral Consolidado", f"='Planilha Orçamentária 3 Etapas'!K{tot_r}", "Teto Conforme Edital", "✓ 100% Dentro do Teto Solicitado", currency_fmt),
                    ("Custos Administrativos (Teto 15%)", f"='Planilha Orçamentária 3 Etapas'!I6", "Teto Máximo 15% (IN MinC)", "✓ Conforme (<= 15%)", currency_fmt),
                    ("Comunicação & Divulgação (Teto 10%)", f"='Planilha Orçamentária 3 Etapas'!I8", "Teto Máximo 10% (Fomento)", "✓ Conforme (<= 10%)", currency_fmt),
                    ("Acessibilidade PCD Obrigatória", "Intérprete LIBRAS + Audiodescrição", "Obrigatório (Lei 13.146/15)", "✓ Atendido Integralmente", None),
                    ("Encargos & Tributos (ISS/INSS)", f"='Planilha Orçamentária 3 Etapas'!J{tot_r}", "Retenções Fiscais na Fonte", "✓ Provisionado no Orçamento", currency_fmt),
                ]
                
                for idx, (ind, val_f, teto, par, nfmt) in enumerate(sum_rows):
                    r = 4 + idx
                    ws3.cell(row=r, column=1, value=ind).font = bold_font
                    c_val = ws3.cell(row=r, column=2, value=val_f)
                    c_val.font = bold_font
                    if nfmt: c_val.number_format = nfmt
                    ws3.cell(row=r, column=3, value=teto).font = regular_font
                    ws3.cell(row=r, column=4, value=par).font = bold_font
                    for c_idx in range(1, 5):
                        ws3.cell(row=r, column=c_idx).border = thin_border

                for col in ws3.columns:
                    max_l = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws3.column_dimensions[col_letter].width = max(max_l + 3, 16)

                # ABA 4: Análise de Tetos & Conformidade
                validacao_tetos = data.get('validacaoTetos', [])
                if validacao_tetos:
                    ws4 = wb.create_sheet(title="Análise de Tetos")
                    ws4.cell(row=1, column=1, value="Análise de Tetos & Conformidade Orçamentária").font = title_font
                    ws4.cell(row=2, column=1, value=f"Dashboard de conformidade para: {project_title}").font = sub_font
                    
                    headers4 = ["Grupo / Rubrica", "Valor Total (R$)", "% do Orçamento", "Teto Permitido", "Status de Conformidade"]
                    for c_idx, h in enumerate(headers4, start=1):
                        c = ws4.cell(row=4, column=c_idx, value=h)
                        c.fill = summary_header_fill; c.font = header_font; c.border = thin_border
                    
                    for idx, vt in enumerate(validacao_tetos):
                        r = 5 + idx
                        ws4.cell(row=r, column=1, value=vt.get('grupoRubrica', '')).font = bold_font
                        c_val = ws4.cell(row=r, column=2, value=parse_num(vt.get('valorTotal', 0)))
                        c_val.font = bold_font; c_val.number_format = currency_fmt
                        c_pct = ws4.cell(row=r, column=3, value=parse_num(vt.get('percentualDoOrcamento', 0)) / 100)
                        c_pct.font = regular_font; c_pct.number_format = pct_fmt
                        ws4.cell(row=r, column=4, value=vt.get('tetoPermitido', 'Sem teto')).font = regular_font
                        ws4.cell(row=r, column=5, value=vt.get('statusConformidade', '✓ Conforme')).font = bold_font
                        for c_idx in range(1, 6):
                            ws4.cell(row=r, column=c_idx).border = thin_border
                    
                    for col in ws4.columns:
                        max_l = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws4.column_dimensions[col_letter].width = max(max_l + 3, 14)

                # ABA 5: Encargos Trabalhistas & Tributários
                regime_items = [it for it in items if it.get('regimeTributario') and it.get('regimeTributario') not in ('N/A', 'Isento')]
                if regime_items:
                    ws5 = wb.create_sheet(title="Encargos Trabalhistas")
                    ws5.cell(row=1, column=1, value="Detalhamento de Encargos Trabalhistas & Tributários").font = title_font
                    ws5.cell(row=2, column=1, value=f"Breakdown ISS/INSS/IRRF por profissional para: {project_title}").font = sub_font
                    
                    headers5 = ["Item / Profissional", "Regime", "Subtotal (R$)", "ISS", "INSS", "IRRF", "Total Encargos", "Custo Total"]
                    for c_idx, h in enumerate(headers5, start=1):
                        c = ws5.cell(row=4, column=c_idx, value=h)
                        c.fill = timeline_header_fill; c.font = header_font; c.border = thin_border
                    
                    for idx, it in enumerate(regime_items):
                        r = 5 + idx
                        sub = parse_num(it.get('subtotal', 0))
                        regime = it.get('regimeTributario', 'N/A')
                        iss = sub * 0.05 if regime in ('RPA', 'PJ') else 0
                        inss = sub * 0.11 if regime == 'RPA' else (sub * 0.20 if regime == 'CLT' else 0)
                        irrf = sub * 0.075 if regime == 'CLT' else 0
                        
                        row_data = [
                            (clean_str(it.get('item', '')), bold_font),
                            (regime, regular_font),
                            (sub, regular_font),
                            (round(iss), regular_font),
                            (round(inss), regular_font),
                            (round(irrf), regular_font),
                            (round(iss + inss + irrf), bold_font),
                            (round(sub + iss + inss + irrf), bold_font)
                        ]
                        for c_idx, (val, fn) in enumerate(row_data, start=1):
                            c = ws5.cell(row=r, column=c_idx, value=val)
                            c.font = fn; c.border = thin_border
                            if c_idx >= 3: c.number_format = currency_fmt
                    
                    for col in ws5.columns:
                        max_l = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws5.column_dimensions[col_letter].width = max(max_l + 3, 14)

                # ABA 6: Cronograma de Desembolso Mensal
                ws6 = wb.create_sheet(title="Cronograma Desembolso Mensal")
                ws6.cell(row=1, column=1, value="Cronograma Financeiro de Desembolso Mensal").font = title_font
                ws6.cell(row=2, column=1, value=f"Fluxo de caixa para: {project_title}").font = sub_font
                
                cron_mensal = data.get('cronogramaDesembolsoMensal', [])
                if cron_mensal:
                    headers6 = ["Mês", "Fase do Projeto", "Desembolso Previsto (R$)", "Principais Atividades"]
                    for c_idx, h in enumerate(headers6, start=1):
                        c = ws6.cell(row=4, column=c_idx, value=h)
                        c.fill = timeline_header_fill; c.font = header_font; c.border = thin_border
                    
                    total_desembolso = 0
                    for idx, cm in enumerate(cron_mensal):
                        r = 5 + idx
                        val = parse_num(cm.get('valorDesembolso', 0))
                        total_desembolso += val
                        ws6.cell(row=r, column=1, value=f"Mês {cm.get('mes', idx+1)}").font = bold_font
                        ws6.cell(row=r, column=2, value=cm.get('fase', '')).font = regular_font
                        c_val = ws6.cell(row=r, column=3, value=val)
                        c_val.font = bold_font; c_val.number_format = currency_fmt
                        ws6.cell(row=r, column=4, value=cm.get('principaisAtividades', '')).font = regular_font
                        for c_idx in range(1, 5):
                            ws6.cell(row=r, column=c_idx).border = thin_border
                    
                    # Total row
                    tr = 5 + len(cron_mensal)
                    ws6.cell(row=tr, column=1, value="TOTAL").font = bold_font
                    c_tot = ws6.cell(row=tr, column=3, value=total_desembolso)
                    c_tot.font = bold_font; c_tot.number_format = currency_fmt
                    ws6.cell(row=tr, column=4, value="✓ Fluxo Consolidado").font = bold_font
                    for c_idx in range(1, 5):
                        c = ws6.cell(row=tr, column=c_idx)
                        c.fill = total_fill; c.border = thin_border
                else:
                    # Fallback: 3 fases
                    headers6f = ["Fase", "Período", "% Participação", "Desembolso (R$)", "Atividades"]
                    for c_idx, h in enumerate(headers6f, start=1):
                        c = ws6.cell(row=4, column=c_idx, value=h)
                        c.fill = timeline_header_fill; c.font = header_font; c.border = thin_border
                    
                    flow_rows = [
                        ("Fase 1: Pré-Produção", "Mês 1-2", 0.25, f"='Planilha Orçamentária 3 Etapas'!K{tot_r}*0.25", "Contratações e mobilização."),
                        ("Fase 2: Execução", "Mês 3-10", 0.60, f"='Planilha Orçamentária 3 Etapas'!K{tot_r}*0.60", "Atividades principais."),
                        ("Fase 3: Pós-Produção", "Mês 11-12", 0.15, f"='Planilha Orçamentária 3 Etapas'!K{tot_r}*0.15", "Prestação de contas."),
                    ]
                    for idx, (fase, per, pct, form, ativ) in enumerate(flow_rows):
                        r = 5 + idx
                        ws6.cell(row=r, column=1, value=fase).font = bold_font
                        ws6.cell(row=r, column=2, value=per).font = regular_font
                        c_pct = ws6.cell(row=r, column=3, value=pct)
                        c_pct.font = regular_font; c_pct.number_format = pct_fmt
                        c_val = ws6.cell(row=r, column=4, value=form)
                        c_val.font = bold_font; c_val.number_format = currency_fmt
                        ws6.cell(row=r, column=5, value=ativ).font = regular_font
                        for c_idx in range(1, 6):
                            ws6.cell(row=r, column=c_idx).border = thin_border
                
                for col in ws6.columns:
                    max_l = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    ws6.column_dimensions[col_letter].width = max(max_l + 3, 14)

                # ABA 7: Despesas Vedadas (Checklist)
                desp_vedadas = data.get('despesasVedadasChecklist', [])
                if desp_vedadas:
                    ws7 = wb.create_sheet(title="Despesas Vedadas")
                    ws7.cell(row=1, column=1, value="Checklist de Despesas Vedadas pelo Edital").font = title_font
                    ws7.cell(row=2, column=1, value=f"Verificação de conformidade para: {project_title}").font = sub_font
                    
                    headers7 = ["Despesa Vedada (conforme edital)", "Status", "Observação"]
                    for c_idx, h in enumerate(headers7, start=1):
                        c = ws7.cell(row=4, column=c_idx, value=h)
                        c.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
                        c.font = header_font; c.border = thin_border
                    
                    for idx, dv in enumerate(desp_vedadas):
                        r = 5 + idx
                        ws7.cell(row=r, column=1, value=dv.get('despesaVedada', '')).font = bold_font
                        ws7.cell(row=r, column=2, value=dv.get('status', '✓ OK')).font = bold_font
                        ws7.cell(row=r, column=3, value=dv.get('observacao', '')).font = regular_font
                        for c_idx in range(1, 4):
                            ws7.cell(row=r, column=c_idx).border = thin_border
                    
                    for col in ws7.columns:
                        max_l = max(len(str(cell.value or '')) for cell in col)
                        col_letter = get_column_letter(col[0].column)
                        ws7.column_dimensions[col_letter].width = max(max_l + 3, 14)


                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_bytes = excel_buffer.getvalue()
                excel_buffer.close()

                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                import unicodedata
                filename_clean = ''.join(c for c in unicodedata.normalize('NFD', project_title) if unicodedata.category(c) != 'Mn')
                filename_clean = re.sub(r'[^a-zA-Z0-9]', '_', filename_clean)
                filename_clean = re.sub(r'_+', '_', filename_clean).strip('_')
                if not filename_clean or filename_clean.lower() == 'titulo_do_projeto_cultural':
                    filename_clean = "Projeto_Cultural"
                self.send_header('Content-Disposition', f'attachment; filename="Planilha_Financeira_{filename_clean}.xlsx"')
                self.send_header('Content-Length', str(len(excel_bytes)))
                self.end_headers()
                self.wfile.write(excel_bytes)
                return
            except Exception as e:
                import traceback
                traceback.print_exc()
                self.send_json_response(500, {"error": f"Erro ao gerar XLSX do financeiro: {str(e)}"})
        
        elif self.path == '/api/save-audit-report':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                with open('relatorio_auditoria.json', 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                self.send_json_response(200, {"success": True, "message": "Relatório salvo no backend."})
            except Exception as e:
                self.send_json_response(500, {"error": f"Erro ao salvar relatório no backend: {str(e)}"})

        elif self.path == '/api/load-audit-report':
            try:
                if os.path.exists('relatorio_auditoria.json'):
                    with open('relatorio_auditoria.json', 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    self.send_json_response(200, data)
                else:
                    self.send_json_response(404, {"error": "Relatório não encontrado no backend."})
            except Exception as e:
                self.send_json_response(500, {"error": f"Erro ao carregar relatório: {str(e)}"})

        elif self.path == '/api/analyze-edital-context':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                edital_text = data.get('editalRefText', '')
                annexes = data.get('annexes', [])
                api_key = data.get('api_key', '')
                cover = data.get('cover', {})  # Skill A: Receber dados do proponente para cruzamento
                model = data.get('model') or os.environ.get('GEMINI_DEFAULT_MODEL') or 'gemini-3.5-flash'
                model = data.get('model') or os.environ.get('GEMINI_DEFAULT_MODEL') or 'gemini-3.5-flash'
                
                if not edital_text.strip():
                    self.send_json_response(200, {
                        "fomento": "Não Especificado",
                        "objetivos": "Por favor, faça o upload do regulamento do edital.",
                        "tetos_e_limites": "Nenhum valor informado.",
                        "acessibilidade_e_cotas": "Não especificado.",
                        "prioridades_critérios": "Não mapeado.",
                        "anexos_analisados": "Nenhum anexo fornecido."
                    })
                    return

                annexes_context = "\n---\n".join([
                    f"Anexo: {a.get('name', 'Anexo')}\nConteúdo: {a.get('content', '')[:30000]}"
                    for a in annexes
                ]) if annexes else "Sem anexos adicionais."

                # Skill A: Cruzar perfil do proponente com edital se dados disponíveis
                proponent_context = ""
                if cover and (cover.get('title') or cover.get('proponent') or cover.get('city')):
                    proponent_context = f"""\n\n[DADOS DO PROPONENTE PARA CRUZAMENTO ESTRATÉGICO]:
- Título do Projeto: {cover.get('title', 'Não informado')}
- Proponente: {cover.get('proponent', 'Não informado')}
- Cidade/UF: {cover.get('city', 'Não informado')}
- Orçamento Pretendido: R$ {cover.get('budget', 0)}
- Ano de Execução: {cover.get('year', 'Não informado')}

Para a chave "compatibilidade_estrategica", analise a compatibilidade entre o perfil deste proponente (localização, público-alvo provável, capacidade orçamentária) e os objetivos/elegibilidade do edital. Indique grau de aderência e riscos de inelegibilidade."""

                analyze_prompt = f"""Você é o Auditor-Geral e Analista Estrutural de editais públicos e privados.
Sua missão é analisar minuciosamente o Edital principal e seus Anexos fornecidos abaixo para mapear e extrair o perfil estrutural e as regras de conformidade que devem governar todo e qualquer texto ou proposta gerada para este edital.

[CONTEÚDO DO EDITAL DE REFERÊNCIA]:
{edital_text[:150000]}

[ANEXOS ADICIONAIS]:
{annexes_context}
{proponent_context}

Mapeie e estruture as informações em um objeto JSON contendo exatamente as seguintes chaves:
1. fomento: Nome da lei de incentivo ou linha de fomento identificada no edital.
2. objetivos: Resumo curto e claro do foco temático, objetivos principais do edital e tipos de projetos elegíveis.
3. tetos_e_limites: Valores máximos (teto por projeto) e limites percentuais REAIS para rubricas conforme definidos no edital (ex: limite de custos administrativos, divulgação, assessoria).
4. acessibilidade_e_cotas: Regras obrigatórias de acessibilidade e políticas de ação afirmativa/cotas conforme o edital.
5. prioridades_critérios: Critérios de prioridade, desempate e avaliação extraídos dos anexos de pontuação.
6. anexos_analisados: Lista compacta dos anexos enviados e a importância de cada um para o projeto.
7. secoes_exigidas: Lista contendo apenas as chaves das seções especificamente exigidas ou necessárias conforme o edital e anexos (escolhidas estritamente entre: "justificativa", "objetivos", "metodologia", "cronograma", "orcamento", "acessibilidade", "publico", "contrapartida", "comunicacao", "ficha_tecnica", "monitoramento", "compliance", "sustentabilidade", "rider").
8. compatibilidade_estrategica: Parecer breve sobre a compatibilidade entre o proponente e o edital (grau de aderência, riscos de inelegibilidade, oportunidades). Se nenhum dado do proponente foi fornecido, retorne "Dados do proponente não informados para análise de compatibilidade.".

Retorne estritamente o JSON estruturado conforme o Schema fornecido. Sem blocos markdown ou explicações fora do JSON."""

                ANALYZE_SCHEMA = {
                    "type": "OBJECT",
                    "properties": {
                        "fomento": {"type": "STRING"},
                        "objetivos": {"type": "STRING"},
                        "tetos_e_limites": {"type": "STRING"},
                        "acessibilidade_e_cotas": {"type": "STRING"},
                        "prioridades_critérios": {"type": "STRING"},
                        "anexos_analisados": {"type": "STRING"},
                        "secoes_exigidas": {
                            "type": "ARRAY",
                            "items": {"type": "STRING"}
                        },
                        "compatibilidade_estrategica": {"type": "STRING"}
                    },
                    "required": ["fomento", "objetivos", "tetos_e_limites", "acessibilidade_e_cotas", "prioridades_critérios", "anexos_analisados", "secoes_exigidas", "compatibilidade_estrategica"]
                }

                print("[SERVER] Iniciando análise prévia do edital...")
                result_str = gateway.generate(
                    provider_name='gemini',
                    model=model,
                    api_key=api_key,
                    prompt=analyze_prompt,
                    system_instruction="Você é o analista estrutural de editais. Retorne estritamente um JSON estruturado com o perfil do edital.",
                    response_schema=ANALYZE_SCHEMA,
                    use_cache=True
                )
                
                try:
                    result_json = json.loads(result_str)
                except Exception as e:
                    print(f"[SERVER][ERROR] Erro ao decodificar JSON de análise: {e}")
                    # Fallback parser
                    try:
                        clean_str = result_str.strip()
                        if clean_str.startswith("```json"):
                            clean_str = clean_str[7:]
                        if clean_str.endswith("```"):
                            clean_str = clean_str[:-3]
                        result_json = json.loads(clean_str.strip())
                    except Exception as e2:
                        print(f"[SERVER][ERROR] Segundo parser falhou: {e2}")
                        raise e
                
                self.send_json_response(200, result_json)
            except urllib.error.HTTPError as he:
                import traceback
                traceback.print_exc()
                if he.code == 429:
                    self.send_json_response(429, {"error": "Limite de requisições do Gemini excedido (HTTP 429). Por favor, aguarde alguns instantes antes de tentar novamente ou verifique os limites de sua chave de API."})
                elif he.code == 400:
                    self.send_json_response(400, {"error": "Requisição inválida para a API do Gemini (HTTP 400). Verifique a chave de API ou as regras configuradas."})
                else:
                    self.send_json_response(he.code, {"error": f"Erro na API do Gemini (HTTP {he.code}): {he.reason}"})
            except Exception as e:
                import traceback
                traceback.print_exc()
                if "429" in str(e):
                    self.send_json_response(429, {"error": "Limite de requisições do Gemini excedido (HTTP 429). Por favor, aguarde alguns instantes antes de tentar novamente."})
                else:
                    self.send_json_response(500, {"error": f"Erro na análise do edital: {str(e)}"})

        elif self.path == '/api/generate-proposal-unified':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                cover = data.get('cover', {})
                editalRefText = data.get('editalRefText', '')
                proposalDraftText = data.get('proposalDraftText', '')
                ingestaoNotes = data.get('ingestaoNotes', '')
                annexes = data.get('annexes', [])
                historicalMemories = data.get('historicalMemories', [])
                editalProfile = data.get('editalProfile', {})
                api_key = data.get('api_key', '')
                model = data.get('model') or os.environ.get('GEMINI_DEFAULT_MODEL') or 'gemini-3.5-flash'
                
                profile_context = f"""[PERFIL E REGRAS ESTRUTURAIS DO EDITAL (MANDATÓRIO CRUZAMENTO)]:
- Fomento / Lei: {editalProfile.get('fomento', 'N/A')}
- Objetivos / Elegibilidade: {editalProfile.get('objetivos', 'N/A')}
- Tetos e Limites: {editalProfile.get('tetos_e_limites', 'N/A')}
- Acessibilidade e Cotas: {editalProfile.get('acessibilidade_e_cotas', 'N/A')}
- Prioridades e Critérios: {editalProfile.get('prioridades_critérios', 'N/A')}
- Anexos Mapeados: {editalProfile.get('anexos_analisados', 'N/A')}
""" if editalProfile else ""

                # Pre-processing contexts (generous limits to prevent truncation)
                annexes_context = "\n---\n".join([
                    f"Anexo: {a.get('name', 'Anexo')}\nConteúdo: {a.get('content', '')[:30000]}"
                    for a in annexes
                ]) if annexes else "Sem anexos adicionais."
                
                # Determinar quais seções gerar com base no perfil do edital
                all_sections = ['justificativa', 'objetivos', 'metodologia', 'cronograma', 'orcamento', 'acessibilidade', 'publico', 'contrapartida', 'comunicacao', 'ficha_tecnica', 'monitoramento', 'compliance', 'sustentabilidade', 'rider']
                secoes_exigidas = editalProfile.get('secoes_exigidas', all_sections) if editalProfile else all_sections
                if not secoes_exigidas or len(secoes_exigidas) == 0:
                    secoes_exigidas = all_sections
                
                # Descritores dinâmicos para cada seção
                section_descriptors = {
                    'justificativa': 'Justificativa longa, detalhada e persuasiva defendendo o mérito, relevância social e impacto no território.',
                    'objetivos': 'Objetivo geral claro e objetivos específicos listados como itens de realizações quantificáveis.',
                    'metodologia': 'Metodologia operacional detalhando passo-a-passo as fases de Pré-produção, Execução e Pós-produção.',
                    'cronograma': 'Cronograma formatado obrigatoriamente como tabela HTML (<table>) organizado por meses.',
                    'orcamento': 'Planilha orçamentária como tabela HTML com colunas: Item, Quantidade, Unidade, Valor Unitário (R$), Valor Total (R$). Respeite rigorosamente os tetos percentuais REAIS extraídos do edital (NÃO use valores genéricos como 15% ou 10%).',
                    'acessibilidade': 'Plano de acessibilidade física, atitudinal e sensorial/comunicacional e cotas afirmativas exigidas pelo edital.',
                    'publico': 'Público-Alvo: perfil demográfico, social e etário detalhado dos beneficiários.',
                    'contrapartida': 'Contrapartida Social e Legado duradouro oferecido gratuitamente à comunidade.',
                    'comunicacao': 'Plano de Comunicação e Divulgação nas mídias sociais, imprensa e peças gráficas.',
                    'ficha_tecnica': 'Ficha Técnica com minibios e cargos da equipe principal para atestar a exequibilidade operacional.',
                    'monitoramento': 'Plano de Monitoramento, Avaliação e Indicadores de sucesso (Matriz Lógica).',
                    'compliance': 'Mecanismos de compliance legal, certidões negativas necessárias conforme o edital.',
                    'sustentabilidade': 'Plano de Sustentabilidade e práticas ESG para mitigação de impactos ambientais.',
                    'rider': 'Rider Técnico detalhando necessidades físicas, mapa de palco, som/luz, montagem e logística.'
                }
                
                # Gerar descrições apenas para as seções exigidas pelo edital
                sections_text = "\n".join([
                    f"{i+1}. {sec}: {section_descriptors.get(sec, 'Seção solicitada pelo edital.')}"
                    for i, sec in enumerate(secoes_exigidas)
                ])

                unified_prompt = f"""Você é uma inteligência artificial de elite especialista em captação de recursos públicos e editais de fomento.
Sua missão é realizar um cruzamento exaustivo e rigoroso entre os dados do edital, seus anexos, o rascunho fornecido e as anotações/orientações específicas do proponente para redigir uma proposta completa de altíssimo nível.

**INSTRUÇÕES CRÍTICAS DE REDAÇÃO (EVITE RESPOSTAS GENÉRICAS):**
- Redija cada seção de forma densa, completa, profissional e contextualizada para o projeto. Não faça resumos, resenhas ou redações rasas.
- Incorpore profundamente o conteúdo e as ideias presentes no [RASCUNHO DO PROPONENTE] e em [ANOTAÇÕES E PONTOS DE ATENÇÃO DO PROPONENTE].
- Respeite e atenda estritamente aos tetos financeiros, limites percentuais, regras de acessibilidade e critérios de priorização descritos no [PERFIL E REGRAS ESTRUTURAIS DO EDITAL] e no [CONTEÚDO DO EDITAL].
- A redação deve estar pronta para submissão oficial (sem placeholders como "[inserir nome]", "[definir data]" ou marcas/pistas de IA).

{profile_context}

[DADOS DO PROJETO]:
- Título: {cover.get('title', 'Não informado')}
- Instituição: {cover.get('institution', 'Não informado')}
- Proponente: {cover.get('proponent', 'Não informado')}
- Cidade/UF: {cover.get('city', 'Não informado')}
- Ano: {cover.get('year', 'Não informado')}
- Orçamento Teto do Projeto: R$ {cover.get('budget', 0)}

[CONTEÚDO DO EDITAL DE REFERÊNCIA (Regulamento)]:
{editalRefText[:150000]}

[ANEXOS ADICIONAIS DO EDITAL]:
{annexes_context}

[RASCUNHO DO PROPONENTE]:
{proposalDraftText[:50000]}

[ANOTAÇÕES E PONTOS DE ATENÇÃO DO PROPONENTE (DIRECIONAMENTO DOS AGENTES DE DOMÍNIO)]:
{ingestaoNotes[:20000] if ingestaoNotes else "Nenhuma anotação adicional."}

---

### MISSÃO: REDIGIR AS SEÇÕES DA PROPOSTA CONFORME EXIGIDO PELO EDITAL
Gere a redação das seguintes seções identificadas como obrigatórias/relevantes pelo perfil do edital (deve conter tags HTML de cabeçalho h3 ou h4 e parágrafos dentro de cada texto):
{sections_text}

Retorne estritamente o JSON estruturado conforme o Schema fornecido. Sem trechos em markdown ou explicações fora do JSON."""

                UNIFIED_RESPONSE_SCHEMA = {
                    "type": "OBJECT",
                    "properties": {
                        "documentContent": {
                            "type": "OBJECT",
                            "properties": {
                                "justificativa": {"type": "STRING"},
                                "objetivos": {"type": "STRING"},
                                "metodologia": {"type": "STRING"},
                                "cronograma": {"type": "STRING"},
                                "orcamento": {"type": "STRING"},
                                "acessibilidade": {"type": "STRING"},
                                "publico": {"type": "STRING"},
                                "contrapartida": {"type": "STRING"},
                                "comunicacao": {"type": "STRING"},
                                "ficha_tecnica": {"type": "STRING"},
                                "monitoramento": {"type": "STRING"},
                                "compliance": {"type": "STRING"},
                                "sustentabilidade": {"type": "STRING"},
                                "rider": {"type": "STRING"}
                            },
                            "required": [
                                "justificativa", "objetivos", "metodologia", "cronograma", "orcamento", "acessibilidade",
                                "publico", "contrapartida", "comunicacao", "ficha_tecnica", "monitoramento", "compliance",
                                "sustentabilidade", "rider"
                            ]
                        }
                    },
                    "required": ["documentContent"]
                }

                print("[SERVER] Iniciando geração da proposta em uma única chamada...")
                result_str = gateway.generate(
                    provider_name='gemini',
                    model=model,
                    api_key=api_key,
                    prompt=unified_prompt,
                    system_instruction="Você é o orquestrador especialista de projetos culturais. Retorne estritamente um JSON contendo documentContent.",
                    response_schema=UNIFIED_RESPONSE_SCHEMA,
                    use_cache=False
                )
                
                try:
                    result_json = json.loads(result_str)
                except Exception as e:
                    print(f"[SERVER][ERROR] Erro ao decodificar JSON unificado: {e}")
                    # Fallback parser
                    try:
                        clean_str = result_str.strip()
                        if clean_str.startswith("```json"):
                            clean_str = clean_str[7:]
                        if clean_str.endswith("```"):
                            clean_str = clean_str[:-3]
                        result_json = json.loads(clean_str.strip())
                    except Exception as e2:
                        print(f"[SERVER][ERROR] Segundo parser unificado falhou: {e2}")
                        raise e
                
                self.send_json_response(200, result_json)
            except urllib.error.HTTPError as he:
                import traceback
                traceback.print_exc()
                if he.code == 429:
                    self.send_json_response(429, {"error": "Limite de requisições do Gemini excedido (HTTP 429). Por favor, aguarde alguns instantes antes de tentar novamente ou verifique os limites de sua chave de API."})
                elif he.code == 400:
                    self.send_json_response(400, {"error": "Requisição inválida para a API do Gemini (HTTP 400). Verifique a chave de API ou as regras configuradas."})
                else:
                    self.send_json_response(he.code, {"error": f"Erro na API do Gemini (HTTP {he.code}): {he.reason}"})
            except Exception as e:
                import traceback
                traceback.print_exc()
                if "429" in str(e):
                    self.send_json_response(429, {"error": "Limite de requisições do Gemini excedido (HTTP 429). Por favor, aguarde alguns instantes antes de tentar novamente."})
                else:
                    self.send_json_response(500, {"error": f"Erro na geração unificada: {str(e)}"})

        elif self.path == '/api/export-anki':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                deck_name = data.get('deck_name', 'Baralho_Concursos_SRS')
                flashcards = data.get('flashcards', [])
                zip_bytes = create_anki_apkg_zip(deck_name, flashcards)
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/zip')
                self.send_header('Content-Disposition', f'attachment; filename="{deck_name}.apkg"')
                self.send_header('Content-Length', str(len(zip_bytes)))
                self.end_headers()
                self.wfile.write(zip_bytes)
            except Exception as ex:
                self.send_json_response(500, {"error": f"Erro ao gerar pacote Anki: {str(ex)}"})

        elif self.path == '/api/llm/generate':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                provider = 'gemini'
                model = data.get('model') or os.environ.get('GEMINI_DEFAULT_MODEL') or 'gemini-3.5-flash'
                api_key = data.get('api_key', '')
                prompt = data.get('prompt', '')
                system_instruction = data.get('system_instruction', None)
                ollama_url = data.get('ollama_url', None)
                use_cache = data.get('use_cache', True)
                use_chunking = data.get('use_chunking', True)
                stream = data.get('stream', False)
                response_schema = data.get('response_schema', None)
                
                # --- TRUNCAMENTO DE SEGURANÇA NO SERVIDOR ---
                MAX_EDITAL_CHARS = 1000000
                MAX_ANNEX_CHARS = 500000
                MAX_FINAL_PROMPT_CHARS = 600000
                
                # Context items for RAG (with safety truncation)
                edital_text = data.get('edital_text', '')
                if edital_text and len(edital_text) > MAX_EDITAL_CHARS:
                    print(f"[SERVER][WARN] edital_text truncado: {len(edital_text)} -> {MAX_EDITAL_CHARS} chars")
                    edital_text = edital_text[:MAX_EDITAL_CHARS]
                
                annexes = data.get('annexes', [])
                if annexes:
                    for a in annexes:
                        a_content = a.get('content', '')
                        if a_content and len(a_content) > MAX_ANNEX_CHARS:
                            print(f"[SERVER][WARN] Anexo '{a.get('name', '?')}' truncado: {len(a_content)} -> {MAX_ANNEX_CHARS} chars")
                            a['content'] = a_content[:MAX_ANNEX_CHARS]
                
                # If chunking is enabled and we have edital/annexes text, perform retrieval
                if use_chunking:
                    retrieved_context = []
                    
                    compliance_keywords = (
                        "orçamento limite teto custos administrativo tributário imposto taxa RPA INSS MEI ISS "
                        "regularidade certidão FGTS CND CNDT proponente tempo atuação experiência justificativa "
                        "cronograma fase pré-produção execução pós-produção prazo riscos equipe direitos autorais "
                        "ECAD som imagem cessão SisGen acessibilidade física rampa adaptada Libras audiodescrição "
                        "cotas reserva democratização contrapartida oficina workshop formação doação "
                        "prestação contas verificação presença foto nota fiscal comprovante clipagem mídia"
                    )
                    retrieval_query = f"{prompt}\n{compliance_keywords}"
                    
                    if edital_text:
                        if len(edital_text) <= 120000:
                            # Keep 100% of the edital if it's within 120k chars (~30-40 pages)
                            retrieved_context.append("=== CONTEÚDO DO EDITAL DE REFERÊNCIA ===\n" + edital_text)
                        else:
                            # Use generous RAG for larger editais
                            edital_chunks = DocumentRetriever.retrieve(edital_text, retrieval_query, top_k=15)
                            if edital_chunks:
                                retrieved_context.append("=== TRECHOS RELEVANTES DO EDITAL ===\n" + "\n---\n".join(edital_chunks))
                                
                    if annexes:
                        annex_chunks_list = []
                        for a in annexes:
                            a_name = a.get('name', 'Anexo')
                            a_content = a.get('content', '')
                            if not a_content:
                                continue
                            if len(a_content) <= 4000:
                                # Keep small annexes fully
                                annex_chunks_list.append(f"Anexo: {a_name}\n{a_content}")
                            else:
                                # Use RAG for larger annexes
                                chunks = DocumentRetriever.retrieve(a_content, retrieval_query, top_k=3)
                                if chunks:
                                    annex_chunks_list.append(f"Anexo: {a_name}\n" + "\n---\n".join(chunks))
                        if annex_chunks_list:
                            retrieved_context.append("=== TRECHOS RELEVANTES DOS ANEXOS EXTRAS ===\n" + "\n---\n".join(annex_chunks_list))
                            
                    if retrieved_context:
                        context_str = "\n\n".join(retrieved_context)
                        prompt = f"{prompt}\n\n[CONTEXTO RELEVANTE RECUPERADO (RAG)]:\n{context_str}"
                
                # --- TRUNCAMENTO FINAL DO PROMPT ---
                if len(prompt) > MAX_FINAL_PROMPT_CHARS:
                    print(f"[SERVER][WARN] Prompt final truncado: {len(prompt)} -> {MAX_FINAL_PROMPT_CHARS} chars")
                    prompt = prompt[:MAX_FINAL_PROMPT_CHARS]
                
                # --- LOG DE DEPURAÇÃO ---
                print(f"[SERVER][DEBUG] /api/llm/generate | Modelo: {model} | Prompt final: {len(prompt)} chars | Cache: {use_cache}")
                
                # Call the gateway
                if stream:
                    self.send_response(200)
                    self.send_header('Content-Type', 'text/event-stream')
                    self.send_header('Cache-Control', 'no-cache')
                    self.send_header('Connection', 'keep-alive')
                    self.end_headers()
                    try:
                        for chunk in gateway.stream_generate(
                            provider_name=provider,
                            model=model,
                            api_key=api_key,
                            prompt=prompt,
                            system_instruction=system_instruction,
                            ollama_url=ollama_url,
                            use_cache=use_cache,
                            response_schema=response_schema
                        ):
                            event_data = f"data: {json.dumps({'text': chunk})}\n\n"
                            self.wfile.write(event_data.encode('utf-8'))
                            self.wfile.flush()
                        self.wfile.write(b"data: [DONE]\n\n")
                        self.wfile.flush()
                    except Exception as stream_err:
                        print(f"[SERVER][STREAM][ERROR] {stream_err}")
                        error_data = f"data: {json.dumps({'error': str(stream_err)})}\n\n"
                        self.wfile.write(error_data.encode('utf-8'))
                        self.wfile.write(b"data: [DONE]\n\n")
                        self.wfile.flush()
                    return
                else:
                    gateway_response = gateway.generate(
                        provider_name=provider,
                        model=model,
                        api_key=api_key,
                        prompt=prompt,
                        system_instruction=system_instruction,
                        ollama_url=ollama_url,
                        use_cache=use_cache,
                        response_schema=response_schema
                    )
                    self.send_json_response(200, {"text": gateway_response})
            except TimeoutError as e:
                print(f"[SERVER][TIMEOUT] {str(e)}")
                self.send_json_response(504, {"error": f"Timeout: {str(e)}"})
            except urllib.error.HTTPError as e:
                import traceback
                traceback.print_exc()
                try:
                    error_body = e.read().decode('utf-8')
                except:
                    error_body = str(e)
                if e.code == 429:
                    self.send_json_response(429, {"error": "Limite de requisições do Gemini excedido (HTTP 429). Por favor, aguarde alguns instantes."})
                elif e.code == 400:
                    self.send_json_response(400, {"error": "Requisição inválida para a API do Gemini (HTTP 400). Verifique a chave de API ou as regras configuradas."})
                else:
                    self.send_json_response(500, {"error": f"Erro na API do Provedor (HTTP {e.code}): {error_body}"})
            except Exception as e:
                import traceback
                traceback.print_exc()
                if "429" in str(e):
                    self.send_json_response(429, {"error": "Limite de requisições do Gemini excedido (HTTP 429). Por favor, aguarde alguns instantes."})
                else:
                    self.send_json_response(500, {"error": f"Erro no LLM Gateway: {str(e)}"})
        
        else:
            self.send_json_response(404, {"error": "Rota de API não encontrada."})

    def send_json_response(self, status_code, data_dict):
        self.send_response(status_code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        response_bytes = json.dumps(data_dict).encode('utf-8')
        self.send_header('Content-Length', str(len(response_bytes)))
        self.end_headers()
        self.wfile.write(response_bytes)


def main():
    # Garante que serve a pasta atual (onde index.html está localizado)
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
    server_address = ('127.0.0.1', PORT)
    httpd = ThreadingHTTPServer(server_address, CustomHTTPRequestHandler)
    print(f"Servidor EditalAudit AI rodando em http://127.0.0.1:{PORT}/")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor encerrado.")

if __name__ == '__main__':
    main()
